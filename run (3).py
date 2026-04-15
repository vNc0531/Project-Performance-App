import os
import json
import uuid
import datetime as dt
import pandas as pd
import re
import openpyxl

from flask import Flask, render_template, request, jsonify
from werkzeug.utils import secure_filename
from sqlalchemy import create_engine, text
app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# =========================================================
# DATABASE CONNECTION
# =========================================================
# Put these in your .env or export them in terminal:
# MSSQL_SERVER=...
# MSSQL_DB=ProjectPerformance
# MSSQL_USER=...
# MSSQL_PASSWORD=...

#MSSQL_SERVER = os.getenv("MSSQL_SERVER", r"NSJC0290\SQLEXPRESS")
#MSSQL_DB = os.getenv("MSSQL_DB", "ProjectPerformance")

#CONN_STR = (
#       f"mssql+pyodbc://@{MSSQL_SERVER}/{MSSQL_DB}"
#        "?driver=ODBC+Driver+18+for+SQL+Server"
#      "&TrustServerCertificate=yes"
#)
#
#engine = create_engine(CONN_STR, fast_executemany=True)
#with engine.connect() as conn:
#    result = conn.execute(text("Select 1"))
#    print("DB test:", result.scalar())

# =========================================================
# HELPERS
# =========================================================
def safe_json_loads(s: str):
    if not s:
        return []
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        return []

def read_table_file(path):
    """
    Read CSV or Excel into pandas DataFrame.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        return pd.read_csv(path)
    elif ext in [".xlsx", ".xls"]:
        return pd.read_excel(path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def clean_columns(df):
    """
    Strip spaces from column names.
    """
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    return df

# -----------------------------
# To Tool Ready
# -----------------------------

TOOLREADY_COLUMNS = {"Expense Category", "Original hourly estimates (hours)"}

ALLOWED_ESTIMATE_CATEGORIES = {
    "PjM",
    "SA",
    "RFENGR",
    "RFENGMGR",
    "RFTECH",
    "ME",
    "EE",
    "SWMGR",
    "DVEMGR",
    "HWENGMGR",
    "NPIMFGMGR",
    "NPI test",
    "NPIMFG",
    "REL",
    "QE",
    "SQE",
}

CATEGORY_ALIASES = {
    "NPIMFG ": "NPIMFG",
    "REL ": "REL",
    "EE  ": "EE",
    "EE": "EE",
    "SQE ": "SQE",
    "QE ": "QE",
}

SECTION_ROWS_TO_SKIP = {
    "Engineering Labor",
    "Non Engineering Labor",
}

STOP_AT_ROWS = {
    "Contract Work",
    "Travel",
    "Project Materials/Other",
    "R&D Labor",
}

def normalize_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()

def coerce_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip().replace(",", "")
    if s in {"", "-", "None"}:
        return None

    try:
        return float(s)
    except ValueError:
        return None

def looks_like_toolready_estimate_file(path: str) -> bool:
    try:
        df = pd.read_excel(path, nrows=5)
        cols = {str(c).strip() for c in df.columns}
        return TOOLREADY_COLUMNS.issubset(cols)
    except Exception:
        return False

def extract_toolready_estimate_df_from_original(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Summary" not in wb.sheetnames:
        raise ValueError(f"{os.path.basename(path)} has no 'Summary' sheet.")

    ws = wb["Summary"]

    header_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        col_b = normalize_text(ws.cell(r, 2).value).lower()
        col_d = normalize_text(ws.cell(r, 4).value).lower()
        if col_b == "expense category" and "original hourly estimates" in col_d:
            header_row = r
            break

    if header_row is None:
        raise ValueError("Could not find Summary header row.")

    records = []

    for r in range(header_row + 1, ws.max_row + 1):
        category = normalize_text(ws.cell(r, 2).value)
        hours = ws.cell(r, 4).value

        if category in STOP_AT_ROWS:
            break
        if not category or category in SECTION_ROWS_TO_SKIP:
            continue

        category = CATEGORY_ALIASES.get(category, category)
        hours = coerce_number(hours)

        if category in ALLOWED_ESTIMATE_CATEGORIES and hours is not None and hours >= 0:
            records.append({
                "Expense Category": category,
                "Original hourly estimates (hours)": hours,
            })

    if not records:
        raise ValueError("No estimate rows extracted from Summary sheet.")

    return pd.DataFrame(records)

def ensure_toolready_estimate_file(path: str, output_dir: str | None = None) -> str:
    if looks_like_toolready_estimate_file(path):
        return path

    df = extract_toolready_estimate_df_from_original(path)

    if output_dir is None:
        output_dir = os.path.dirname(path) or "."

    base = os.path.splitext(os.path.basename(path))[0]
    out_path = os.path.join(output_dir, f"{base}__toolready_generated.xlsx")
    df.to_excel(out_path, index=False)
    return out_path

def build_combined_teamhours(project_id, est_path, cado_path, map_path):
    """
    Build one combined DataFrame for a single project:
    ProjectID, CostCenter, FunctionGroup, Estimated_Hours, Actual_Hours, UpdatedAt
    """

    # -----------------------------
    # Read files
    # -----------------------------
    est_path = ensure_toolready_estimate_file(est_path, output_dir=UPLOAD_FOLDER)
    est_df = clean_columns(read_table_file(est_path))
    cado_df = clean_columns(read_table_file(cado_path))
    map_df = clean_columns(read_table_file(map_path))

    # -----------------------------
    # Standardize mapping columns
    # -----------------------------
    map_df = map_df.rename(columns={
        "ExpenseCategory": "ExpenseCategory",
        "Expense Category": "ExpenseCategory",
        "SenderCostCenter": "SenderCostCenter",
        "Sender Cost Center": "SenderCostCenter",
        "FunctionGroup": "FunctionGroup",
        "Function Group": "FunctionGroup",
        "Name": "Name",
        "Employee Name": "Name",
        "EmployeeName": "Name",
    }).copy()

    for col in ["ExpenseCategory", "SenderCostCenter", "FunctionGroup", "Name"]:
        if col not in map_df.columns:
            map_df[col] = ""

    map_df["ExpenseCategory"] = map_df["ExpenseCategory"].fillna("").astype(str).str.strip()
    map_df["SenderCostCenter"] = map_df["SenderCostCenter"].fillna("").astype(str).str.strip()
    map_df["FunctionGroup"] = map_df["FunctionGroup"].fillna("").astype(str).str.strip()
    map_df["Name"] = map_df["Name"].fillna("").astype(str).str.strip()
    map_df["NameNorm"] = map_df["Name"].str.lower()

    # -----------------------------
    # Build mapping dictionaries
    # -----------------------------
    # generic cost center -> function group
    generic_cc_rows = map_df[
        map_df["SenderCostCenter"].ne("") & map_df["Name"].eq("")
    ][["SenderCostCenter", "FunctionGroup"]].drop_duplicates()

    cc_to_team = dict(generic_cc_rows.values)

    # override: cost center + name -> function group
    override_rows = map_df[
        map_df["SenderCostCenter"].ne("") & map_df["Name"].ne("")
    ][["SenderCostCenter", "NameNorm", "FunctionGroup"]].drop_duplicates()

    cc_name_to_team = {
        f"{row['SenderCostCenter']}||{row['NameNorm']}": row["FunctionGroup"]
        for _, row in override_rows.iterrows()
    }

    # expense category -> function group
    exp_rows = map_df[
        map_df["ExpenseCategory"].ne("")
    ][["ExpenseCategory", "FunctionGroup"]].drop_duplicates()

    exp_rows["ExpenseCategoryNorm"] = exp_rows["ExpenseCategory"].str.lower()
    exp_to_team = dict(exp_rows[["ExpenseCategoryNorm", "FunctionGroup"]].values)

    # function group -> default cost center
    # for estimates / reporting
    fg_to_cc = {}

    # use generic mapping rows first
    generic_fg_cc_rows = map_df[
        map_df["SenderCostCenter"].ne("") & map_df["Name"].eq("")
    ][["FunctionGroup", "SenderCostCenter"]].drop_duplicates()

    for _, row in generic_fg_cc_rows.iterrows():
        fg = row["FunctionGroup"]
        cc = row["SenderCostCenter"]
        # keep first seen mapping
        if fg and fg not in fg_to_cc:
            fg_to_cc[fg] = cc

    # force RF and HW to both use 24020
    fg_to_cc["RF"] = "24020"
    fg_to_cc["HW"] = "24020"

    # -----------------------------
    # Estimated side
    # -----------------------------
    est_df = est_df.rename(columns={
        "Expense Category": "ExpenseCategory",
        "Original hourly estimates (hours)": "Estimated_Hours",
    }).copy()

    if "ExpenseCategory" not in est_df.columns:
        est_df["ExpenseCategory"] = ""
    if "Estimated_Hours" not in est_df.columns:
        est_df["Estimated_Hours"] = 0

    est_df["ExpenseCategory"] = est_df["ExpenseCategory"].fillna("").astype(str).str.strip()
    est_df["ExpenseCategoryNorm"] = est_df["ExpenseCategory"].str.lower()
    est_df["Estimated_Hours"] = pd.to_numeric(est_df["Estimated_Hours"], errors="coerce").fillna(0)

    est_df["FunctionGroup"] = est_df["ExpenseCategoryNorm"].map(exp_to_team)

    # Map each expense category to its cost center from the mapping file
    exp_to_cc = dict(
        zip(
            map_df["ExpenseCategory"].str.lower(),
            map_df["SenderCostCenter"].astype(str),
        )
    )
    est_df["CostCenter"] = est_df["ExpenseCategoryNorm"].map(exp_to_cc).fillna("")

    # Force RF and HW estimates to report under 24020
    est_df.loc[est_df["FunctionGroup"].isin({"RF", "HW"}), "CostCenter"] = "24020"

    est_grouped = (
        est_df.dropna(subset=["FunctionGroup"])
        .groupby(["FunctionGroup", "CostCenter"], as_index=False)["Estimated_Hours"]
        .sum()
    )
    est_grouped["ProjectID"] = project_id

    # -----------------------------
    # Actual side (CADO)
    # -----------------------------
    cado_df = cado_df.rename(columns={
        "Sender Cost Center": "SenderCostCenter",
        "Hours": "Actual_Hours",
        "Name of employee or applicant": "EmployeeName",
        "Employee Name": "EmployeeName",
    }).copy()

    if "SenderCostCenter" not in cado_df.columns:
        cado_df["SenderCostCenter"] = ""
    if "Actual_Hours" not in cado_df.columns:
        cado_df["Actual_Hours"] = 0

    if "EmployeeName" not in cado_df.columns:
        if "Name" in cado_df.columns:
            cado_df["EmployeeName"] = cado_df["Name"]
        else:
            cado_df["EmployeeName"] = ""

    if isinstance(cado_df["EmployeeName"], pd.DataFrame):
        cado_df["EmployeeName"] = cado_df["EmployeeName"].iloc[:, 0]

    cado_df["SenderCostCenter"] = cado_df["SenderCostCenter"].fillna("").astype(str).str.strip()
    cado_df["Actual_Hours"] = pd.to_numeric(cado_df["Actual_Hours"], errors="coerce").fillna(0)
    cado_df["EmployeeName"] = cado_df["EmployeeName"].fillna("").astype(str).str.strip()
    cado_df["EmployeeNameNorm"] = cado_df["EmployeeName"].str.lower()

    def map_cado_row(row):
        cc = row["SenderCostCenter"]
        nm = row["EmployeeNameNorm"]

        if not cc:
            return None

        # Special rule for shared 24020:
        # named people are HW, everyone else under 24020 is RF
        if cc == "24020":
            if nm:
                key = f"{cc}||{nm}"
                if key in cc_name_to_team:
                    return cc_name_to_team[key]   # should hit HW named overrides
            return "RF"

        # all other cost centers follow normal mapping
        if nm:
            key = f"{cc}||{nm}"
            if key in cc_name_to_team:
                return cc_name_to_team[key]

        return cc_to_team.get(cc, None)

    cado_df["FunctionGroup"] = cado_df.apply(map_cado_row, axis=1)

    # final reporting cost center:
    # RF and HW both should be 24020
    def reporting_cost_center(row):
        fg = row["FunctionGroup"]
        cc = row["SenderCostCenter"]

        if fg in {"RF", "HW"}:
            return "24020"

        return cc

    cado_df["CostCenter"] = cado_df.apply(reporting_cost_center, axis=1)

    actual_grouped = (
        cado_df.dropna(subset=["FunctionGroup"])
        .groupby(["CostCenter", "FunctionGroup"], as_index=False)["Actual_Hours"]
        .sum()
    )
    actual_grouped["ProjectID"] = project_id

    # -----------------------------
    # Combine estimated + actual
    # -----------------------------
    combined = actual_grouped.merge(
        est_grouped[["ProjectID", "FunctionGroup", "CostCenter", "Estimated_Hours"]],
        on=["ProjectID", "FunctionGroup", "CostCenter"],
        how="outer",
    )

    combined["Estimated_Hours"] = combined["Estimated_Hours"].fillna(0)
    combined["Actual_Hours"] = combined["Actual_Hours"].fillna(0)
    combined["CostCenter"] = combined["CostCenter"].fillna("")

    combined["UpdatedAt"] = dt.datetime.now(dt.timezone.utc)

    combined = combined[
        ["ProjectID", "CostCenter", "FunctionGroup", "Estimated_Hours", "Actual_Hours", "UpdatedAt"]
    ]

    return combined

def sql_value(v):
    return None if pd.isna(v) else v

def quote_ident(name: str) -> str:
    return f"[{str(name).replace(']', ']]')}]"

def replace_project_rows_then_append(df, table_name, project_col="ProjectID", schema="core"):
    """
    Delete all existing rows for the incoming project IDs, then append the new rows.
    This is safer than key-based replacement when business logic changes.
    """
    if df is None or df.empty:
        return {
            "table": f"{schema}.{table_name}",
            "mode": "replace_project_rows_then_append",
            "rows_incoming": 0,
            "rows_appended": 0,
        }

    df = df.copy()

    if project_col not in df.columns:
        raise ValueError(f"{project_col} not found in dataframe")

    df[project_col] = df[project_col].astype(str).str.strip()
    df = df[df[project_col] != ""]

    if df.empty:
        return {
            "table": f"{schema}.{table_name}",
            "mode": "replace_project_rows_then_append",
            "rows_incoming": 0,
            "rows_appended": 0,
        }

    project_ids = df[project_col].dropna().astype(str).str.strip().unique().tolist()

    q_table = f"{quote_ident(schema)}.{quote_ident(table_name)}"
    delete_sql = text(
        f"DELETE FROM {q_table} WHERE {quote_ident(project_col)} = :project_id"
    )

#    with engine.begin() as conn:
#        for pid in project_ids:
#            conn.execute(delete_sql, {"project_id": pid})
#
#       df.to_sql(
#            table_name,
#            con=conn,
#            schema=schema,
#            if_exists="append",
#            index=False,
#            method="multi",
#            chunksize=1000,
#        )

    return {
        "table": f"{schema}.{table_name}",
        "mode": "replace_project_rows_then_append",
        "project_ids_replaced": project_ids,
        "rows_incoming": int(len(df)),
        "rows_appended": int(len(df)),
    }

def delete_matching_rows_then_append(df, table_name, key_cols, schema="core"):
    """
    For each incoming row:
      1) delete existing DB row with same key
      2) append the new row
    Leaves all other rows untouched.
    """
    if df is None or df.empty:
        return {
            "table": f"{schema}.{table_name}",
            "mode": "delete_matching_rows_then_append",
            "rows_incoming": 0,
            "rows_appended": 0
        }

    df = df.copy()

    # clean string keys
    for col in key_cols:
        if col in df.columns and df[col].dtype == object:
            df[col] = df[col].astype(str).str.strip()

    if "ProjectID" in df.columns:
        df["ProjectID"] = df["ProjectID"].astype(str).str.strip()

    # remove bad key rows
    for col in key_cols:
        df = df[df[col].notna()]
        if df[col].dtype == object:
            df = df[df[col].astype(str).str.strip() != ""]

    if df.empty:
        return {
            "table": f"{schema}.{table_name}",
            "mode": "delete_matching_rows_then_append",
            "rows_incoming": 0,
            "rows_appended": 0
        }

    # if same key appears multiple times in upload, keep the last one
    df = df.drop_duplicates(subset=key_cols, keep="last")

    q_table = f"{quote_ident(schema)}.{quote_ident(table_name)}"
    where_clause = " AND ".join(
        f"{quote_ident(col)} = :{col}" for col in key_cols
    )

    delete_sql = text(f"""
        DELETE FROM {q_table}
        WHERE {where_clause}
    """)

#    with engine.begin() as conn:
#        # delete only matching rows
#        for _, row in df[key_cols].iterrows():
#            params = {col: sql_value(row[col]) for col in key_cols}
#            conn.execute(delete_sql, params)

        # append the new rows
#        df.to_sql(
#            table_name,
#            con=conn,
#            schema=schema,
#            if_exists="append",
#            index=False,
#            method="multi",
#           chunksize=1000
#    )

    return {
        "table": f"{schema}.{table_name}",
        "mode": "delete_matching_rows_then_append",
        "key_cols": key_cols,
        "rows_incoming": int(len(df)),
        "rows_appended": int(len(df))
    }

@app.get("/")
def index():
    return render_template("form.html")


@app.get("/health")
def health():
    return jsonify({"status": "ok"})


@app.post("/process")
def process():
    load_id = str(uuid.uuid4())
    loaded_at = dt.datetime.utcnow().isoformat()

    # Manual fields
    project_class = request.form.get("project_class", "").strip()

    tg_general = safe_json_loads(request.form.get("tg_general_json", "[]"))
    tg_costs = safe_json_loads(request.form.get("tg_costs_json", "[]"))
    scope_rows = safe_json_loads(request.form.get("scope_rows_json", "[]"))

    saved_files = []

    # -----------------------------
    # Save team mapping file
    # -----------------------------
    team_map_path = None

    team_map_file = request.files.get("team_map_file")
    if team_map_file and team_map_file.filename:
        fn = secure_filename(team_map_file.filename)
        team_map_path = os.path.join(UPLOAD_FOLDER, f"{load_id}__team_map__{fn}")
        team_map_file.save(team_map_path)
        saved_files.append({"type": "team_map", "filename": fn})

    # -----------------------------
    # Save project hours files
    # -----------------------------
    pids = request.form.getlist("ph_projectid[]")
    est_files = request.files.getlist("ph_est[]")
    cado_files = request.files.getlist("ph_cado[]")

    project_rows = []

    for i, raw_pid in enumerate(pids):
        pid = (raw_pid or "").strip()
        if not pid:
            continue

        est_path = None
        cado_path = None

        if i < len(est_files) and est_files[i] and est_files[i].filename:
            fn = secure_filename(est_files[i].filename)
            est_path = os.path.join(UPLOAD_FOLDER, f"{load_id}__{pid}__est__{fn}")
            est_files[i].save(est_path)
            saved_files.append({"type": "est", "project_id": pid, "filename": fn})

        if i < len(cado_files) and cado_files[i] and cado_files[i].filename:
            fn = secure_filename(cado_files[i].filename)
            cado_path = os.path.join(UPLOAD_FOLDER, f"{load_id}__{pid}__cado__{fn}")
            cado_files[i].save(cado_path)
            saved_files.append({"type": "cado", "project_id": pid, "filename": fn})

        # only keep complete project upload rows
        if pid and est_path and cado_path:
            project_rows.append({
                "ProjectID": pid,
                "est_path": est_path,
                "cado_path": cado_path
            })

    # -----------------------------
    # General Data Table
    # -----------------------------

    tg_general_df = pd.DataFrame(tg_general)

    tg_general_df = tg_general_df.rename(columns={
        "TG3ActualDate": "TG3_Actual_Date",
        "TG3EstimatedDate": "TG3_Estimated_Date",
        "TG1Date": "TG1_Date"})

    if not tg_general_df.empty and "ProjectID" in tg_general_df.columns:
        tg_general_df["ProjectID"] = tg_general_df["ProjectID"].astype(str).str.strip()

    if not tg_general_df.empty:
        tg_general_df["UpdatedAt"] = dt.datetime.now(dt.timezone.utc)

    # -----------------------------
    # Estimated & Actual Hour Table
    # -----------------------------

    all_combined = []

    if team_map_path:
        for row in project_rows:
            combined_df = build_combined_teamhours(
                project_id=row["ProjectID"],
                est_path=row["est_path"],
                cado_path=row["cado_path"],
                map_path=team_map_path
            )
            all_combined.append(combined_df)

    if all_combined:
        final_combined_df = pd.concat(all_combined, ignore_index=True)
    else:
        final_combined_df = pd.DataFrame()

    # -----------------------------
    # Labor Hour Summary Table
    # -----------------------------

    if not final_combined_df.empty:
        labor_hours_df = (
        final_combined_df
        .groupby("ProjectID", as_index=False)[["Estimated_Hours", "Actual_Hours"]]
        .sum()
    )

        labor_hours_df["Variance_Hours"] = (
            labor_hours_df["Actual_Hours"] - labor_hours_df["Estimated_Hours"]
    )   

        labor_hours_df["Variance_Percent"] = (
            labor_hours_df["Variance_Hours"] / labor_hours_df["Estimated_Hours"].replace(0, pd.NA)
    )

        labor_hours_df["UpdatedAt"] = dt.datetime.now(dt.timezone.utc)
    else:
        labor_hours_df = pd.DataFrame()


    # -----------------------------
    # Cost Breakdown Table
    # -----------------------------

    tg_costs_df = pd.DataFrame(tg_costs)

    tg_cost_cols = [
        "ProjectID",
        "TG1_MaterialCost",
        "TG1_DevCost_NoNRE",
        "TG1_NPICost",
        "TG1_TotalBudget",
        "TG3_MaterialCost",
        "TG3_DevCost_NoNRE",
        "TG3_NPICost",
        "TG3_TotalCost",
        "TG4_MaterialCost",
        "TG4_DevCost_NoNRE",
        "TG4_NPICost",
        "TG4_TotalCost",
    ]
    tg_costs_df = tg_costs_df.reindex(columns=tg_cost_cols)

    if not tg_costs_df.empty:
        tg_costs_df = tg_costs_df.rename(columns={
        "TG1_MaterialCost": "TG1_Material_Cost",
        "TG1_DevCost_NoNRE": "TG1_DevCost_NoNRE",
        "TG1_NPICost": "TG1_NPI_Cost",
        "TG1_TotalBudget": "TG1_Forecasted_Budget",
        "TG3_MaterialCost": "TG3_Material_Cost",
        "TG3_DevCost_NoNRE": "TG3_DevCost_NoNRE",
        "TG3_NPICost": "TG3_NPI_Cost",
        "TG3_TotalCost": "TG3_Actual_Cost",
        "TG4_MaterialCost": "TG4_Material_Cost",
        "TG4_DevCost_NoNRE": "TG4_DevCost_NoNRE",
        "TG4_NPICost": "TG4_NPI_Cost",
        "TG4_TotalCost": "TG4_Actual_Cost",
    })    

    if not tg_costs_df.empty and "ProjectID" in tg_costs_df.columns:
        tg_costs_df["ProjectID"] = tg_costs_df["ProjectID"].astype(str).str.strip()

    if not tg_costs_df.empty:
        tg_costs_df["UpdatedAt"] = dt.datetime.now(dt.timezone.utc)
    
    
    # -----------------------------
    # Scope Cluster & Project Similarity Table
    # -----------------------------

    tg_scope_df = pd.DataFrame(scope_rows)

    if not tg_scope_df.empty and "ProjectID" in tg_scope_df.columns:
        tg_scope_df["ProjectID"] = tg_scope_df["ProjectID"].astype(str).str.strip()

    if not tg_scope_df.empty:
        tg_scope_df["UpdatedAt"] = dt.datetime.now(dt.timezone.utc)

    db_result = {}

# 1) General / Project Master table
    if not tg_general_df.empty:
        db_result["project_master"] = delete_matching_rows_then_append(
        tg_general_df,
        table_name="Project_Master",
        key_cols=["ProjectID"],
        schema="core"
    )

# 2) Labor Hour Summary table
    if not labor_hours_df.empty:
        db_result["labor_hours"] = delete_matching_rows_then_append(
        labor_hours_df,
        table_name="Fact_LaborHours",
        key_cols=["ProjectID"],
        schema="core"
    )

# 3) Hours by Team table
    if not final_combined_df.empty:
        db_result["team_hours"] = replace_project_rows_then_append(
        final_combined_df,
        table_name="Fact_TeamHours",
        project_col="ProjectID",
        schema="core"
    )

# 4) Costs table
    if not tg_costs_df.empty:
        db_result["costs"] = delete_matching_rows_then_append(
        tg_costs_df,
        table_name="Fact_Costs",
        key_cols=["ProjectID"],
        schema="core"
    )

# 5) Scope Clusters table
    if not tg_scope_df.empty:
        db_result["scope_clusters"] = delete_matching_rows_then_append(
        tg_scope_df,
        table_name="Fact_ScopeClusters",
        key_cols=["ProjectID"],
        schema="core"
    )
    # -----------------------------
    # Save transformed table to temp Excel
    # -----------------------------
    debug_excel_path = None

    if not final_combined_df.empty:
        debug_excel_path = os.path.join(UPLOAD_FOLDER, f"{load_id}__debug_teamhours.xlsx")
        final_combined_df.to_excel(debug_excel_path, index=False)

    # Return summary so you can see it worked
    return jsonify({
        "ok": True,
        "load_id": load_id,
        "loaded_at_utc": loaded_at,
        "project_class": project_class,
        "counts": {
            "tg_general_rows": len(tg_general),
            "tg_costs_rows": len(tg_costs),
            "scope_rows": len(scope_rows),
            "project_ids": len([p for p in pids if p.strip()]),
            "projects_with_complete_hours_files": len(project_rows),
            "combined_rows": int(len(final_combined_df))
        },
        "saved_files": saved_files,
        "debug_excel_path": debug_excel_path,
        "combined_preview": (
            final_combined_df.head(20).to_dict(orient="records")
            if not final_combined_df.empty else []
        ),
        "tg_general_preview": (tg_general_df.to_dict(orient="records") 
        if not tg_general_df.empty else []
        ),
        "tg_costs_preview": (tg_costs_df.to_dict(orient="records") 
        if not tg_costs_df.empty else []
        ),
        "tg_scope_clusters_preview": (tg_scope_df.to_dict(orient="records") 
        if not tg_scope_df.empty else []
        ),
        "labor_hours_preview": (labor_hours_df.to_dict(orient="records")
        if not labor_hours_df.empty else []
        ),
    })

if __name__ == "__main__":
        app.run(port=5001, debug=True)